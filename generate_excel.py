from openpyxl import Workbook
from openpyxl.styles import Alignment


def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Details"

    title_row = 'Employee ID', 'Employee Full Name', 'User ID', 'Password', 'Department', 'Email', 'Salary', \
        'Job Position', ' Designation', \
        'Manager', 'Date of Joining', 'Current Experience', 'Date of Birth', 'Gender', 'Phone No', 'Bank Name',\
        'Account Holder Name', 'Account Number', 'IFSC Code', \
        'Aadhar Card Number', 'PAN Card No', 'Passport No', 'PF Account No', \
        'UAN No', 'ESIC No', 'CL', 'PL', 'SL', 'LWP', 'Basic', 'HRA', 'DA', 'Other Allowance', 'Incentive', 'Arrears',\
        'Grs Outstanding Adjustment', 'Statutory Bonus', 'Gross Salary', 'EPFO', 'PT', 'TDS', 'Leave Deduction', \
        'Other Deduction', 'Ded Outstanding Adjustment', 'Net Salary', 'Home Loan Application No', 'Home Loan Amount',\
        'Home Loan Person Name', 'Home Loan Interest', 'Period of Home Loan', 'Premium Insurance Number',\
        'Premium Annual Amount', 'Premium Person Name', 'Health Insurance (Self) No',\
        'Health Insurance (Self) Person Name', 'Health Insurance (Self) Annual Amount',\
        'Health Insurance (Self) Period', 'Health Insurance (Spouse) No', 'Health Insurance (Spouse) Person Name',\
        'Health Insurance (Spouse) Annual Amount', 'Health Insurance (Spouse) Period',\
        'Health Insurance (Father) No', 'Health Insurance (Father) person Name', \
        'Health Insurance (Father) Annual Amount', 'Health Insurance (Father) Period', 'Interest on Home Loan Payable',\
        'Interest of Home loan Person Name', 'interest of Home Loan PAN of Lender', \
        'Annual House Rent of Current Month', 'Rent House Landloard Name', 'Rent House Landloard PAN',\
        'Rent House Landloard Address', 'Rent House Period', 'ELSS Annual Amount', 'ELSS Period',\
        'Tution Fees Annual Amount', 'Tution Fees Period'

    ws.append(title_row)

    alignment = Alignment(horizontal='center')
    for n in range(1, len(title_row) + 1):
        ws.column_dimensions[ws.cell(row=2, column=n).column_letter].width = 30
        for cell in ws[str(n)]:
            cell.alignment = alignment


    return wb

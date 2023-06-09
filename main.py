import datetime
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, request, redirect, url_for, session, send_file, Response
import io
from leave_manage import Leavemanage
from firebase_admin import credentials
from firebase_admin import firestore
from details import Profile
from create_new_employee import Create
from salary_manage import Salarymanage
from department import Department
from update_employee import Update_information
from dashboard import Dashboard
import re
from excel_sheet import SalaryData
import concurrent.futures
from salary_slip import SalarySlip
from register import Register
from login import Login
from moth_days import MonthCount
from mail import Mail
import concurrent.futures
from salary_calculation import SalaryCalculation
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from generate_excel import create_excel_file
from read_data import ExcelData
from apscheduler.schedulers.background import BackgroundScheduler
import os
import platform


# FLASK APP
app = Flask(__name__)
app.secret_key = 'tO$&!|0wkamvVia0?n$NqIRVWOG'
app.config['SESSION_TYPE'] = 'filesystem'
# USE A SERVICE ACCOUNT
cred = credentials.Certificate('empoyee-payroll-system-firebase-adminsdk-5h89d-1602329ca8.json')
db = firestore.client()
leaveobj = Leavemanage(db)
dept = Department(db)
update_obj = Update_information(db)
dashboard_obj = Dashboard(db)
register_obj = Register(db)
login_obj = Login(db)
moth_count = MonthCount()
mail_obj = Mail()
companyname='alian_software'
# Testing path
# C:/Users/alian/Desktop/Testing
# C:/Users/alian/Downloads/my_file.xlsx


def clear_session_data():
    with app.app_context():
        session.clear()

scheduler = BackgroundScheduler()
scheduler.add_job(clear_session_data, 'interval', days=1, start_date=datetime.datetime.now().replace(hour=0, minute=0, second=0))
scheduler.start()






@app.route('/', methods=["POST", "GET"])
def login():
    global companyname
    responce = ''
    if request.method == 'POST':
        data = request.form
        result = {}
        for key, value in data.items():
            result[key] = value
        responce = login_obj.login(data, companyname)

        if responce == 'Admin':
            return redirect(url_for('employee_list', username=responce))
        elif responce != False:
            if responce['type'] == 'HR':
                return redirect(url_for('employee_list', username=responce['name']))
            else:
                return redirect(url_for('employee_view', username=responce['name'],
                                        id=responce['empid']))
        else:
            responce = 'Inavalid Id and Password'
    company_list = []
    collections = db.collections()
    for collection in collections:
        company_list.append(collection.id)
    ''' LOGIN PAGE '''
    url = f'/'
    return render_template('login.html', responce=responce, url=url, company_list=company_list)

@app.route('/forgot_password', methods=["POST", "GET"])
def forgot_password():
    if request.method == 'POST':
        print(request.form.to_dict())
        email = request.form.to_dict()['email']
        if email != None:
            Admin = db.collection(companyname).document('admin').get().to_dict()
            if Admin != None:
                company_mail = Admin['AdminID']
                auth_password = Admin['auth_password']
                if company_mail == email:
                    email = email
                    password = Admin['password']
                    mail_obj.forgot_mail(email=email, password=password,
                                         company_mail=company_mail, auth_password=auth_password)
                else:
                    user = db.collection(companyname).document('employee').collection('employee').where(
                        'email', "==", email).get()
                    if user != None:
                        data = user[0].to_dict()
                        email = email
                        password = data['password']
                        mail_obj.forgot_mail(email=email, password=password,
                                             company_mail=company_mail, auth_password=auth_password)

    return redirect(url_for('login'))

@app.route('/success', methods=["POST", "GET"])
def success():
    return render_template('success.html')


# @app.route('/', methods=['GET', 'POST'])
# def register():
#     responce = ''
#     if request.method == 'POST':
#         data = request.form
#         companyname = data["CompanyName"]
#         responce = register_obj.register(data)
#         companyname=companyname
#
#         if responce == True:
#             return redirect(url_for('login', ))
#
#     ''' REGISTER PAGE '''
#     return render_template('register.html', responce=responce)


@app.route('/<username>/dashboard', methods=['GET', 'POST'])
def dashboard(username):
    session.pop('excel_path', default=None)

    if request.method == 'POST':
        form = request.form.to_dict()

        data = {str(form["date"]): form["description"]}
        db.collection(companyname).document('holidays').update(data)
        holidays = db.collection(companyname).document('holidays').get().to_dict()
        moath_data = moth_count.count(holidays)
        db.collection(companyname).document('month_data').set(moath_data)

    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count(holidays)
    working_days = moath_data['workingDays']
    # Check the current date
    if datetime.datetime.now().day == 1 and 'session_leave' not in session:
        leaveobj.leave_add(companyname)
        session['session_leave'] = datetime.datetime.now()
        print("create session Variable")


    if datetime.datetime.today().day == 1 and datetime.datetime.today().month == 1:

        leaveobj.leave_reset(companyname)

    ''' DISPLAY DASHBOARD '''
    dashboard_data = dashboard_obj.all_data(companyname)
    holidays = moth_count.get_holidays(holidays)

    employee_data = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        users_ref = db.collection(companyname).document('employee').collection('employee')
        for emp_doc in users_ref.stream():
            employee_data.append(executor.submit(dashboard_obj._get_employee_data, emp_doc))
    employee_on_leave, total_leaves, employee_birthday, employee_anniversary = {}, {}, {}, {}
    for future in concurrent.futures.as_completed(employee_data):
        result = future.result()
        if 'birthday' in result:
            employee_birthday[result['name']] = result['birthday']

            # employee_birthday = sorted(employee_birthday.items(), key=lambda x: x[1], reverse=True)
        if 'anniversary' in result:
            employee_anniversary[result['name']] = result['anniversary']

        if result['leaves']:
            employee_on_leave[result['name']] = result['leaves']

        if result['total_leaves'] > 0:
            total_leaves[result['name']] = result['total_leaves']

    total_leaves = sorted(total_leaves.items(), key=lambda x: x[1], reverse=True)
    total_leaves = dict(total_leaves[:5])
    print(employee_on_leave)
    # Get current month and year

    # Print the resulting dictionary

    return render_template('dashboard.html', employee_on_leave=employee_on_leave, total_leaves=total_leaves,
                           employee_birthday=employee_birthday, employee_anniversary=employee_anniversary,
                           moath_data=moath_data, holidays=holidays, username=username, dashboard_data=dashboard_data)


@app.route('/<username>/employeelist', methods=['GET', 'POST'])
def employee_list(username):
    if 'increment' not in session:
        increments=db.collection('alian_software').document('increments').get().to_dict()['increments']
        for increment in increments:
            converted_date = datetime.datetime.strptime(increment['effectiveDate'], '%Y-%m-%d').date()
            if datetime.datetime.today().date()==  converted_date:
                print(increment['empid'],int(increment['total']))
                db.collection('alian_software').document('employee').collection('employee').document(increment['empid']).update({'salary': int(increment['total']) * 12})
        session['increment'] = 'Done'
    # SENDING EMPLOYEE MAIL FOR ADD DETAILS
    if request.method == 'POST':
        employee_mail = request.form.get('new_email')
        auth_data = db.collection(companyname).document('admin').get().to_dict()
        company_mail = auth_data['AdminID']
        auth_password = auth_data['auth_password']
        mail_obj.new_employee_mail(email=employee_mail, company_mail=company_mail, auth_password=auth_password)

    ''' DISPLAY LIST OF EMPLOYEES IN COMPANY '''

    # GET ALL EMPLOYEES DETAILS
    def get_employee_data():

        docs = db.collection(companyname).document(u'employee').collection('employee').stream()
        employee_list = {}
        for doc in docs:
            if 'user_status' not in doc.to_dict():
                employee_list.update({doc.id: doc.to_dict()})
            elif doc.to_dict()['user_status'] != 'disable':
                employee_list.update({doc.id: doc.to_dict()})
        return employee_list

    # GET DEPARTMENTS IN ORGANIZAT  ION
    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        employee_data = executor.submit(get_employee_data)
        department_data = executor.submit(get_department_data)
    employee_list = employee_data.result()
    department = department_data.result()

    return render_template('employees_list.html', data=employee_list, department=department, username=username)


@app.route("/<username>/storage-path", methods=["POST"])
def excel_sheet_path(username):
    excel_path = request.json["excel_path"]
    excel = ExcelData(db)
    excel.store_excel_data(companyname, excel_path)
    return redirect(url_for('dashboard', username=username))


@app.route('/<username>/result', methods=['POST', 'GET'])
def add(username):

    ''' NEW EMPLOYEE DATA STORE IN DATABASE AND DISPLAY IN LIST '''
    create = Create(db, companyname)
    # create.result()
    employee_mail = request.form.get('email')
    password = request.form.get('password')
    new_id = create.result()
    auth_data = db.collection(companyname).document('admin').get().to_dict()
    company_mail = auth_data['AdminID']
    auth_password = auth_data['auth_password']
    mail_obj.employee_registered_mail(email=employee_mail, password=password, company_mail=company_mail, auth_password=auth_password, new_id=new_id)

    return redirect(url_for('employee_list',username=username))



@app.route('/<username>/<id>/delete', methods=['POST', 'GET'])
def delete_employee(username, id):
    doc_ref = db.collection(companyname).document(u'employee').collection('employee').document(id)
    doc_ref.update({'user_status': 'disable'})
    return redirect(url_for('employee_list', username=username))


@app.route('/register_employee', methods=['POST', 'GET'])
def employee_register_by_mail():

    # GET ALL EMPLOYEES DETAILS
    def get_employee_data():

        docs = db.collection(companyname).document(u'employee').collection('employee').stream()
        employee_list = {}
        for doc in docs:
            if 'user_status' not in doc.to_dict():
                employee_list.update({doc.id: doc.to_dict()})
            elif doc.to_dict()['user_status'] != 'disable':
                employee_list.update({doc.id: doc.to_dict()})
        return employee_list

    data = get_employee_data()

    responce = ''
    if request.method == 'POST':
        data = request.form
        email = data["email"]

        doc = db.collection(companyname).document("employee").collection('employee').where('email', "==", email).get()
        if len(doc) > 0:
            responce = 'Already Email Exist Try Different Email'
        else:

            ''' NEW EMPLOYEE DATA STORE IN DATABASE AND DISPLAY IN LIST '''
            create = Create(db, companyname)
            create.result()
            employee_mail = request.form.get('email')
            auth_data = db.collection(companyname).document('admin').get().to_dict()
            company_mail = auth_data['AdminID']
            auth_password = auth_data['auth_password']
            mail_obj.employee_registered_mail(employee_mail, companyname, company_mail, auth_password)

            return redirect(url_for('login'))

    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        department_data = executor.submit(get_department_data)
    department = get_department_data()
    return render_template('add_employee.html',department=department,
                           department_data=department_data, responce=responce, data=data)

@app.route('/create_excel')
def create_excel():
    ''' EXCEL SHEET DATA FORMATE FOR NEW COMPANY '''
    # Create the Excel file
    wb = create_excel_file()

    # Save the file to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the file as a response with appropriate headers
    return Response(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment;filename=my_file.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )


@app.route('/upload_file', methods=['POST'])
def upload_file(companyname):
    if request.method == 'POST':
        file = request.files['file']
        if file.filename != '':
            workbook = load_workbook(filename=file)
            worksheet = workbook.active
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)
            return redirect(url_for('employee_list', ))
    return redirect(url_for('employee_list', ))


@app.route('/<username>/employeeprofile/<id>', methods=['GET', 'POST'])
def employee_profile(username, id):
    ''' DISPLAY EMPLOYEE DETAILS '''
    users_ref = db.collection(str(companyname)).document('employee').collection('employee').document(id).collection(
        'leaveMST')
    if request.method == 'POST':
        ''' Store leave Data '''
        result = request.form

        with ThreadPoolExecutor(max_workers=2) as executor:
            executor.submit(leaveobj.take_leave, users_ref, data=result)

    ''' GET LEAVE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        total_leave_future = executor.submit(leaveobj.get_total_leave, users_ref)
        leave_list_future = executor.submit(leaveobj.leave_list, users_ref)

    total_leave = total_leave_future.result()
    leave_list = leave_list_future.result()

    ''' GET EMPLOYEE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        personal_data_future = executor.submit(Profile(db, id, companyname).personal_data)
        tds_data_future = executor.submit(Profile(db, id, companyname).tds_data)
        salary_data_future = executor.submit(Profile(db, id, companyname).salary_data)

    increment_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})
    increment_list = [(index, content) for index, content in enumerate(increment_data)]

    contract_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('contract'):
            contract_data.append({key: value})
    contract_list = [(index, content) for index, content in enumerate(contract_data)]

    data = {'personal_data': personal_data_future.result(), 'tds_data': tds_data_future.result(),
            'salary_data': salary_data_future.result()}

    leave_status = False
    leave_status_date = ''
    for i in leave_list:
        if i != 'total_leaves':
            leave_date = (datetime.datetime.strptime(leave_list[i]['applydate'], '%Y-%m-%d').month)
            if leave_date == datetime.datetime.today().month:
                leave_status = True
                leave_status_date = (f'{leave_list[i]["fromdate"]} to {leave_list[i]["todate"]} ')

    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        department_data = executor.submit(get_department_data)
    department = department_data.result()
    return render_template('employee_profile.html', leave=leave_status, data=data, total_leave=total_leave,
                           leave_list=leave_list, leave_date=leave_status_date,username=username, department=department,
                           increment_data=increment_list, contract_data=contract_list)


@app.route('/<username>/employee_view/<id>', methods=['GET', 'POST'])
def employee_view(username, id):
    ''' DISPLAY EMPLOYEE DETAILS '''
    users_ref = db.collection(str(companyname)).document('employee').collection('employee').document(id).collection(
        'leaveMST')
    if request.method == 'POST':
        ''' Store leave Data '''
        result = request.form

        with ThreadPoolExecutor(max_workers=2) as executor:
            executor.submit(leaveobj.take_leave, users_ref, data=result)

    ''' GET LEAVE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        total_leave_future = executor.submit(leaveobj.get_total_leave, users_ref)
        leave_list_future = executor.submit(leaveobj.leave_list, users_ref)

    total_leave = total_leave_future.result()
    leave_list = leave_list_future.result()

    ''' GET EMPLOYEE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        personal_data_future = executor.submit(Profile(db, id, companyname).personal_data)
        tds_data_future = executor.submit(Profile(db, id, companyname).tds_data)
        salary_data_future = executor.submit(Profile(db, id, companyname).salary_data)

    increment_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})
    increment_list = [(index, content) for index, content in enumerate(increment_data)]

    contract_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('contract'):
            contract_data.append({key: value})
    contract_list = [(index, content) for index, content in enumerate(contract_data)]

    data = {'personal_data': personal_data_future.result(), 'tds_data': tds_data_future.result(),
            'salary_data': salary_data_future.result()}

    leave_status = False
    leave_status_date = ''
    for i in leave_list:
        if i != 'total_leaves':
            leave_date = (datetime.datetime.strptime(leave_list[i]['applydate'], '%Y-%m-%d').month)
            if leave_date == datetime.datetime.today().month:
                leave_status = True
                leave_status_date = (f'{leave_list[i]["fromdate"]} to {leave_list[i]["todate"]} ')

    return render_template('employee_view.html', leave=leave_status, data=data, total_leave=total_leave,
                           leave_list=leave_list, leave_date=leave_status_date,username=username,
                           increment_data=increment_list, contract_data=contract_list)


@app.route('/<username>/pdf/<id>/<salid>')
def pdf_personal(username, id, salid):
    ''' SALARY SLIP PDF GENERATION '''
    path = get_download_folder()
    salary = SalarySlip(db)
    responce=salary.salary_slip_personal(companyname, id, salid, path)
    # CHECK THE USER
    return responce


@app.route('/<username>/personal_data_update/<id>', methods=['GET', 'POST'])
def personal_data_update(username, id):
    """ UPDATE EMPLOYEE PERSONAL DETAILS """
    if request.method == 'POST':
        form = request.get_json()
        update_obj.update_personal_info(companyname, form, id)
    return redirect(url_for('employee_profile', id=id, username=username))



""" Update/Add Data in Contract And Increment """
@app.route('/save_data/<empid>/<username>', methods=['POST'])
def save_data(empid, username):
    """ Update/Add Data in Contract And Increment """
    ''' GET EMPLOYEE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        personal_data_future = executor.submit(Profile(db, empid, companyname).personal_data)

    # # INCREMENT DATA
    increment_data = []
    personal_data = personal_data_future.result()
    for key, value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})

    increment_list = [(index, content) for index, content in enumerate(increment_data)]



    # # CONTRACT DATA
    contract_data = []
    personal_data = personal_data_future.result()
    for key, value in personal_data.items():
        if key.startswith('contract'):
            contract_data.append({key: value})
    contract_list = [(index, content) for index, content in enumerate(contract_data)]

    if request.method == 'POST':
        data = request.form.to_dict()

        contract_increment_data = {}
        for n in range(1, 10):
            for key, value in data.items():

                if key.startswith(f'increment_0{n}'):
                    if data[f'increment_0{n}_increment'].endswith('%'):
                        data[f'increment_0{n}_increment'] = str(float(data[f'increment_0{n}_grossSalary']) * float((data[f'increment_0{n}_increment']).split('%')[0]) * 0.01)
                        new_data = {f'increment_0{n}':
                            {
                                f'grossSalary': data[f'increment_0{n}_grossSalary'],
                                f'jobPosition': data[f'increment_0{n}_jobPosition'],
                                f'effectiveDate': data[f'increment_0{n}_effectiveDate'],
                                f'increment': data[f'increment_0{n}_increment'],
                                f'total': data[f'increment_0{n}_total'],
                                f'note': data[f'increment_0{n}_note']}
                        }
                        contract_increment_data.update(new_data)
                    else:
                        new_data = {f'increment_0{n}':
                            {
                                f'grossSalary': data[f'increment_0{n}_grossSalary'],
                                f'jobPosition': data[f'increment_0{n}_jobPosition'],
                                f'effectiveDate': data[f'increment_0{n}_effectiveDate'],
                                f'increment': data[f'increment_0{n}_increment'],
                                f'total': data[f'increment_0{n}_total'],
                                f'note': data[f'increment_0{n}_note']}
                        }


                        contract_increment_data.update(new_data)

                elif key.startswith(f'new_inc_'):
                    if data[f'new_inc_grossSalary'] != "" and data[f'new_inc_jobPosition'] != "":
                        if data[f'new_inc_increment'].endswith('%'):
                            data[f'new_inc_increment'] = str(float(data[f'new_inc_grossSalary']) * float((data[f'new_inc_increment']).split('%')[0]) * 0.01)
                        new_data = {f'increment_0{len(increment_list) + 1}':
                            {
                                f'grossSalary': data[f'new_inc_grossSalary'],
                                f'jobPosition': data[f'new_inc_jobPosition'],
                                f'effectiveDate': data[f'new_inc_effectiveDate'],
                                f'increment': data[f'new_inc_increment'],
                                f'total': data[f'new_inc_total'],
                                f'note': data[f'new_inc_note']}
                        }


                        contract_increment_data.update(new_data)

                elif key.startswith(f'contract_0{n}'):
                    new_data = {f'contract_0{n}':
                        {
                            f'contractDate': data[f'contract_0{n}_contractDate'],
                            f'contractPeriod': data[f'contract_0{n}_contractPeriod'],
                            f'nextContractDate': data[f'contract_0{n}_nextContractDate']
                        }
                    }
                    contract_increment_data.update(new_data)

                elif key.startswith(f'new_ctr_'):
                    new_data = {f'contract_0{len(contract_list) + 1}':
                        {
                            f'contractDate': data[f'new_ctr_contractDate'],
                            f'contractPeriod': data[f'new_ctr_contractPeriod'],
                            f'nextContractDate': data[f'new_ctr_nextContractDate']
                        }
                    }
                    contract_increment_data.update(new_data)

                else:
                    pass



        db.collection(companyname).document('employee').collection('employee').document(empid).update(contract_increment_data)
        if "increment_01" in contract_increment_data and len(contract_increment_data) >len(increment_data):
            inc_key = list(contract_increment_data.keys())
            print(inc_key[-1])

            contract_increment_data[inc_key[-1]]['empid'] = empid

            doc_ref = db.collection(companyname).document('increments')
            data = contract_increment_data[inc_key[-1]]
            print(data)

            doc_ref.update({'increments': firestore.ArrayUnion([data])})

    return redirect(url_for('employee_profile', id=empid, username=username))


@app.route('/<username>/tds_data_update/<id>', methods=['GET', 'POST'])
def tds_data_update(username, id):
    """ UPDATE EMPLOYEE TDS DETAILS """
    if request.method == 'POST':
        form = request.get_json()
        update_obj.update_tds_info(companyname, form, id)
    return redirect(url_for('employee_profile', id=id, username=username))


@app.route('/<username>/department', methods=['GET', 'POST'])
def department(username):
    """ DISPLAY DEPARTMENT """
    if request.method == 'POST':
        # Add New DEPARTMENT
        result = request.form
        dept.add_department(companyname, result)
    doc_ref = db.collection(str(companyname)).document(u'department')
    data = doc_ref.get().to_dict()
    return render_template('department.html', data=data, obj=dept, username=username)


@app.route('/<username>/delete_department/<dep> <pos>', methods=['GET', 'POST'])
def delete_department(username, dep, pos):
    ''' DELETE DEPARTMENT '''
    a = dep, pos
    pattern = r'[^a-zA-Z\d\s]'
    # Use the re.sub() function to replace all occurrences of the pattern with an empty string
    dep = re.sub(pattern, '', dep)
    pos = re.sub(pattern, '', pos)
    dept.delete_department(companyname, dep, pos)
    return redirect(url_for('department', username=username))


@app.route("/<username>/set-storage-path/<salid>", methods=["POST"])
def set_storage_path(username, salid):
    path = request.json["path"]
    salary = SalarySlip(db)
    salary.salary_slip(companyname, salid, path)
    return redirect(url_for('salary', username=username, salid=salid))



@app.route('/<username>/salary', methods=['GET', 'POST'])
def salary(username):
    ''' DISPLAY SALARY DETAILS OF ALL MONTH IN YEAR '''
    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count(holidays)
    working_days = moath_data['workingDays']
    if datetime.datetime.now().day == 1 and 'session_salary' not in session:
        SalaryCalculation(db, companyname).generate_salary(workingday=working_days)
        leaveobj.leave_add(companyname)
        session['session_salary'] = datetime.datetime.now()
        print("create session Variable")

    if request.method == 'POST':
        form = request.form
        data_dict = {}
        for key, value in form.items():
            if value != '':
                data_dict.update({key: value})
        # ADD Salary Criteria
        docs = db.collection(companyname).document('salary_calc')

        if len(docs.get().to_dict()) == 0:
            db.collection(companyname).document('salary_calc').set(data_dict)
        else:
            db.collection(companyname).document('salary_calc').update(data_dict)

    def get_salary_criteria():
        return db.collection(str(companyname)).document('salary_calc').get().to_dict()

    def get_all_month_salary_data():
        return Salarymanage(db).get_all_month_salary_data(companyname)

    def get_salary_status():
        return db.collection(companyname).document('salary_status').get().to_dict()

    with ThreadPoolExecutor(max_workers=3) as executor:
        salary_criteria_future = executor.submit(get_salary_criteria)
        salary_list_future = executor.submit(get_all_month_salary_data)
        salary_status_future = executor.submit(get_salary_status)

    salary_criteria = salary_criteria_future.result()
    salary_list = salary_list_future.result()
    salary_status = salary_status_future.result()
    year = datetime.datetime.now().year
    return render_template('salary_sheet_month.html', data=salary_list, salary_criteria=salary_criteria
                           , username=username, salary_status=salary_status, year=year)


@app.route('/<username>/salarysheetview/<salid>', methods=['GET', 'POST'])
def salary_sheet_view(username, salid):
    # month = int(salid[5:])

    if request.method == 'POST':
        form = request.form
        fields = {}
        for key, value in form.items():
            fields.update({key: value})

        path = get_download_folder()
        salary_excel = SalaryData(db)
        salary_excel.add_data(salid=salid, fields=fields, path=path, companyname=companyname)

    ''' DISPLAY SALARY DETAILS OF EMPLOYEES IN MONTH '''
    salary_list = Salarymanage(db).get_all_emp_salary_data(companyname, salid)

    salary_status = db.collection(companyname).document('salary_status').get()
    salary_status = salary_status.get(datetime.date(1900, int(salid[4:]), 1).strftime('%B'))

    return render_template('salary_sheet_view.html', data=salary_list, salid=salid,
                           username=username, salary_status=salary_status)


@app.route('/<username>/salarysheetedit/<empid> <salid>', methods=['GET', 'POST'])
def salary_sheet_edit_(username, empid, salid):
    ''' EDIT SALARY DETAILS OF EMPLOYEE IN MONTH '''
    if request.method == 'POST':
        result = request.form
        Salarymanage(db).salary_update(companyname, empid, salid, data=result)
        return redirect(url_for('salary_sheet_view', salid=salid, username=username))

    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count(holidays)
    working_days = moath_data['workingDays']
    employee_salary_data = Salarymanage(db).get_salary_data(companyname, empid, salid)
    salary_percentage = (db.collection(companyname).document('salary_calc').get()).to_dict()
    return render_template('salary_sheet_edit_personal.html', data=employee_salary_data, id=salid
                           , salary_data=salary_percentage, username=username,
                           working_days=working_days)


@app.route('/<username>/set_status/<salid>/<status>')
def set_status(username, salid, status):
    ''' SALARY SLIP PDF GENERATION '''
    month = datetime.date(1900, int(salid[3:]), 1).strftime('%B')
    status = status
    data = {month: status}
    salary_status = db.collection(companyname).document('salary_status').update(data)
    return redirect(url_for('salary_sheet_view', username=username, salid=salid))


import os
import platform


def get_download_folder():
    if os.name == 'nt':  # for Windows
        import winreg
        sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
        downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
            location = winreg.QueryValueEx(key, downloads_guid)[0]
    elif os.name == 'darwin':  # for macOS
        location = os.path.expanduser('~/Downloads')
    else:  # for Linux/Unix
        location = os.path.expanduser('~/Downloads')
    return location

from flask import Flask, make_response


@app.route('/download_pdf')
def download_pdf():
    # Open the PDF file
    with open('sample.pdf', 'rb') as file:
        pdf_content = file.read()

    # Create a response object with the PDF content
    response = make_response(pdf_content)
    response.headers['Content-Disposition'] = 'attachment; filename=example.pdf'
    response.headers['Content-Type'] = 'application/pdf'

    return response




@app.route('/<username>/pdf/<salid>')
def pdf(username, salid):
    ''' SALARY SLIP PDF GENERATION '''
    global companyname

    path = get_download_folder()
    salary = SalarySlip(db)
    salary.salary_slip(companyname, salid, path)
    return redirect(url_for('salary', username=username, salid=salid))


# @app.route('/<username>/excel/<salid>')
# def excel_for_bank(companyname,username, salid):
#     ''' GENERATE EXCELSHEET FOR BANK '''
#     if 'storage_path' in session:
#         path = session["storage_path"]
#         salary_excel = SalaryData(db)
#         salary_excel.add_data(companyname, salid, path)
#     return redirect(url_for('salary_sheet_view', salid=salid, username=username))


# @app.route('/tds/<id>', methods=['GET', 'POST'])
# def tds(companyname, id):
#     ''' DISPLAY TDS DETAILS OF EMPLOYEE '''
#     profile = Profile(companyname, db, id)
#     employee_tds_data = {'personal_data': profile.personal_data(), 'tds_data': profile.tds_data()}
#     return render_template('tds_test.html', data=employee_tds_data, )

@app.route('/<username>/add_data', methods=['POST', 'GET'])
def add_data(username):
    if request.method == 'POST':
        data = request.form

        data = data.to_dict()

        excel_path = data['path']

        excel = ExcelData(db)
        excel.store_excel_data(companyname, excel_path)
        return redirect(url_for('dashboard', username=username))
    return render_template('add_excel_file.html', username=username)


@app.route('/<username>/send_email/<salid>')
def send_employee_salaryslip(username, salid):
    ''' GENERATE EXCELSHEET FOR BANK '''

    path = get_download_folder()
    auth_data = db.collection(companyname).document('admin').get().to_dict()
    company_mail = auth_data['AdminID']
    auth_password = auth_data['auth_password']
    docs = db.collection(companyname).document(u'employee').collection('employee').stream()
    employee_list = {}
    for doc in docs:
        employee_list.update({doc.id: doc.to_dict()})
    for key, value in employee_list.items():
        data = value

        mail_obj.send_employee_pdf(company_mail=company_mail, data=data, auth_password=auth_password, path=path,
                                   companyname=companyname)
    return redirect(url_for('salary_sheet_view', salid=salid, username=username))


if __name__ == '__main__':
    # app.run(debug=True, port=300)
    app.run(debug=True, host="0.0.0.0", port=3005)

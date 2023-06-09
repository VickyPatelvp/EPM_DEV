
import calendar

import openpyxl
from salary_manage import Salarymanage
import os
from openpyxl.styles import Alignment

class SalaryData():

    def __init__(self,db):
        self.db=db

    def add_data(self,companyname, salid, fields, path):
        mont_in_num = int(salid[5:])
        month = calendar.month_name[mont_in_num]

        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Store Excelsheet
        file_path = f"{path}/Excelsheets/"

        if not os.path.exists(file_path):
            os.makedirs(file_path)

        file_name = f'Salary_{month}.xlsx'

        excel_file = file_path + file_name

        # Select the active worksheet
        worksheet = workbook.active


        # Worksheet Name
        worksheet.title = "Alian Software"

        # Merge cell for heading
        A1 = worksheet.merge_cells('A1:E1')

        # Heading
        worksheet['A1'] = 'Bank Name Associated With Company'

        title_field = fields

        # Set Table Header
        title_row = [title_field['field1'], title_field['field2'], title_field['field3'], title_field['field4'], title_field['field5']]

        worksheet.append(title_row)

        for n in range(1,6):
            worksheet.column_dimensions[worksheet.cell(row=2, column=n).column_letter].width = 25

        workbook.save(excel_file)

        salary_list = Salarymanage(self.db).get_all_emp_salary_data(companyname,salid)
        print(salary_list)
        for i in salary_list:

            empid = salary_list[i]["userID"]

            user_ref = self.db.collection(companyname).document('employee').collection('employee').document(empid)

            data = user_ref.get().to_dict()

            salary_data = user_ref.collection("salaryslips").document(f"{salid}").get().to_dict()

            data = {
                'Employee Name': data["accountHolderName"],
                'Bank Name': data["bankName"],
                'Account Number': data["accountNumber"],
                'IFSC Code': data["ifscCode"],
                'Salary': salary_data["netSalary"]
            }

            employee_data = [data[title_field['field1']], data[title_field['field2']], data[title_field['field3']], data[title_field['field4']], data[title_field['field5']]]

            worksheet.append(employee_data)

            # Save the workbook
            workbook.save(excel_file)

        # Text Alignment For All Cell
        alignment = Alignment(horizontal='center')
        for n in range(1, 50):
            for cell in worksheet[str(n)]:
                cell.alignment = alignment
        workbook.save(excel_file)

